#!/usr/bin/env python3
import openpyxl


def main():
    cisco = juniper = huawei = NULL = 0

    wb = openpyxl.load_workbook('devices.xlsx')
    sheet = wb["Sheet1"]

    for i in range(2, 2948, 1):
        if 'sz.ip.rostelecom.ru' in sheet.cell(row=i, column=2).value:
            if 'cisco' in sheet.cell(row=i, column=5).value:
                cisco += 1
            elif 'juniper' in sheet.cell(row=i, column=5).value:
                juniper += 1
            elif 'huawei' in sheet.cell(row=i, column=5).value:
                huawei += 1
            else:
                NULL += 1
            print(sheet.cell(row=i, column=2).value, '\t', sheet.cell(row=i, column=5).value)

    print("")
    print('cisco:', str(cisco)+';', 'juniper:', str(juniper)+';', 'huawei:', str(huawei)+';',
          'NULL:', str(NULL)+'.')


if __name__ == "__main__":
    main()

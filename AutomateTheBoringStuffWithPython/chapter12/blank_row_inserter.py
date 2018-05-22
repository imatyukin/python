#!/usr/bin/env python3
import sys
import openpyxl

N = sys.argv[1]
M = sys.argv[2]
filename = str(sys.argv[3])

wb = openpyxl.load_workbook(filename)
before_sheet = wb.active
before_sheet.title = 'before'

wb.create_sheet('after')
after_sheet = wb.get_sheet_by_name('after')

nrows = before_sheet.max_row

for i in range(1, nrows+1):
    for idx, cell in enumerate(before_sheet[i]):
        if i < N:
            after_sheet.cell(row=i, column=idx+1).value = cell.value
        else:
            after_sheet.cell(row=i+M, column=idx+1).value = cell.value

wb.save('example.xlsx')

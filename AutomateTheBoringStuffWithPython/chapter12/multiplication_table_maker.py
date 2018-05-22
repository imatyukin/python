#!/usr/bin/env python3
import sys
import openpyxl
from openpyxl.styles import Font


def main():

    N = int(sys.argv[1])

    wb = openpyxl.Workbook()
    sheet = wb.active

    bold_font = Font(bold=True)

    for row in range(1, N+2):
        for col in range(1, N+2):
            if row == 1 and col == 1:
                continue
            elif row == 1:
                sheet.cell(row=row, column=col).value = col-1
                sheet.cell(row=row, column=col).font = bold_font
            elif col == 1:
                sheet.cell(row=row, column=col).value = row-1
                sheet.cell(row=row, column=col).font = bold_font
            else:
                sheet.cell(row=row, column=col).value = (col-1)*(row-1)

    wb.save('table.xlsx')


if __name__ == "__main__":
    main()

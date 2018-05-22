#!/usr/bin/env python3
import openpyxl


def main():

    filename = "example.xlsx"

    wb = openpyxl.load_workbook(filename)
    before_sheet = wb.active
    before_sheet.title = 'before'

    wb.create_sheet('after')
    after_sheet = wb.get_sheet_by_name('after')

    nrows = before_sheet.max_row
    ncols = before_sheet.max_column

    for row in range(1, nrows+1):
        for col in range(1, ncols+1):
            after_sheet.cell(row=col, column=row).value = before_sheet.cell(row=row, column=col).value

    wb.save('example.xlsx')


if __name__ == "__main__":
    main()

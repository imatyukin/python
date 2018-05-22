#!/usr/bin/env python3
import openpyxl


def main():

    spreadsheet = "text_to_spreadsheet.xlsx"

    wb = openpyxl.load_workbook(spreadsheet)
    sheet = wb.active

    nrows = sheet.max_row
    ncols = sheet.max_column

    for col in range(1, ncols+1):
        text_file = "spreadsheet_" + str(col) + ".txt"
        with open(text_file, 'w') as f:
            for row in range(1, nrows+1):
                content = sheet.cell(row=row, column=col).value
                if content is None:
                    continue
                f.write(str(content))


if __name__ == "__main__":
    main()
